from openpyxl import load_workbook, Workbook  # Import both load and Workbook
from werkzeug.security import generate_password_hash, check_password_hash
from flask import Flask, request, jsonify
from flask_cors import CORS
from flask_cors import cross_origin

app = Flask(__name__)
#CORS(app, resources={r"/*": {"origins": "http://localhost:4200"}})
@app.after_request
def after_request(response):
  response.headers.add('Access-Control-Allow-Origin', '*')
  response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
  response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
  return response


# User modelsss
class User:
    def __init__(self, user_id, name, email, password):
        self.user_id = user_id
        self.name = name
        self.email = email
        self.password = password
        
    # User Registration
    @app.route("/sign-up", methods=["GET"])
    #@cross_origin()
    def register():
        data = request.get_json()
        name = data.get("name")
        email = data.get("email")
        password = data.get("password")
        hashed_password = generate_password_hash(password)
        
        # Open users sheet in write mode (create if not exists)
        wb = load_workbook("users.xlsx", data_only=True)  # Open in read mode initially
    
        # Check if file exists, otherwise create a new one
        if not wb.sheetnames:
            wb = Workbook()
            sheet = wb.active
            sheet.append(["user_id", "name", "email", "password"])  # Add header row

        sheet = wb["users"]  # Access the users sheet
   
        # Check if email already exists
        for row in sheet.iter_rows(values_only=True):  # Read only cell values
            if row[1] == email:
                return jsonify({"message": "Email already exists"}), 400
        
        # Append new user data
        sheet.append([None, name, email, hashed_password])
        wb.save("users.xlsx")

        new_user = User(sheet.max_row, name, email, hashed_password)
    
        return jsonify({"message": "User created successfully", "user": new_user.to_dict()})
        #response.headers.add("Access-Control-Allow-Origin", "*")
        #return response

    # User Login
    @app.route("/login", methods=["POST"])
    def login():
        data = request.get_json()
        email = data.get("email")
        password = data.get("password")

        # Open users sheet in read mode
        wb = load_workbook("users.xlsx", data_only=True)
        sheet = wb["users"]

        # Find user by email
        for row in sheet.iter_rows(values_only=True):
            if row[1] == email:
                user_data = row
                break
            else:
                return jsonify({"message": "Invalid email or password"}), 401
    
        # Verify password
        if not check_password_hash(user_data[3], password):
            return jsonify({"message": "Invalid email or password"}), 401

        return jsonify({"message": "Login successful", "user": {"user_id": user_data[0], "name": user_data[1]}})


# Activity model (updated)
class Activity:
    def __init__(self, activity_id, user_id, timestamp, activity_type, duration, distance, intensity, calories_burned):
        self.activity_id = activity_id
        self.user_id = user_id
        self.timestamp = timestamp
        self.activity_type = activity_type
        self.duration = duration
        self.distance = distance
        self.intensity = intensity
        self.calories_burned = calories_burned

    # Activity Logging
    @app.route("/workout", methods=["POST"])
    def log_activity():
        data = request.get_json()
        #user_id = data.get("user_id")
        #timestamp = data.get("timestamp")
        activity = data.get("activity")
        time = data.get("Time")
        #distance = data.get("distance")
        calories = data.get("calories")
        
        # Open activities sheet in append mode (create if not exists)
        wb = load_workbook("activities.xlsx", data_only=True)  
        if not wb.sheetnames:
            wb = Workbook()
            sheet = wb.active
            sheet.append(["activity", "time","calories"])  # Add header row

        sheet = wb.active  # Access the active sheet
    
        # Append new activity data
        sheet.append([activity, time, calories])
        wb.save("activities.xlsx")

        return jsonify({"message": "Activity logged successfully"})



class Socialshare:
    def __init__(self, activity_id, user_id, timestamp, activity_type, duration, calories_burned):
        self.activity_id = activity_id
        self.user_id = user_id
        self.timestamp = timestamp
        self.activity_type = activity_type
        self.duration = duration
        self.calories_burned = calories_burned

    @app.route("/share/whatsapp", methods=["GET"])
    def share_whatsapp():
        achievement = request.args.get("achievement")  # Get achievement details  (last entry)
        app_link = "https://www.yourapp.com"  # Replace with your app link (optional)
        message = f"🎉 {achievement}! Check out my progress on {app_link}"  # Customize message

        share_url = f"https://api.whatsapp.com/send?text={message}"
        return jsonify({"message": "Share link generated", "share_url": share_url})

if __name__ == "__main__":
    app.run(debug=True)